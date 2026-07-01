"""Excel 分镜表规则解析（无 LLM）。"""

from __future__ import annotations

import re
import uuid
import zipfile
import xml.etree.ElementTree as ET
from typing import Any

HEADER_SCAN_ROWS = 15
HEADER_MIN_KEYWORD_HITS = 2

HEADER_KEYWORDS = ("镜号", "画面", "景别", "时长", "镜头运动", "画面描述")

FIELD_SYNONYMS: dict[str, list[str]] = {
    "shotNumber": ["镜号"],
    "camera": ["景别"],
    "movement": ["镜头运动", "运镜"],
    "duration": ["时长/s", "时长(s)", "时长"],
    "prompt": ["画面描述", "画面"],
    "soundDesign": ["台词/声音", "台词/声音（不唯一）", "台词/声音(不唯一)", "声音", "台词"],
    "composition": ["机位"],
    "atmosphereNote": ["备注"],
}

MAPPABLE_FIELDS = list(FIELD_SYNONYMS.keys()) + ["ignore"]


def _norm_header(text: Any) -> str:
    return re.sub(r"\s+", "", str(text or "").strip())


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def sheet_plaintext(grid: list[list[str]]) -> str:
    lines = []
    for row in grid:
        lines.append("\t".join(_cell_str(c) for c in row))
    return "\n".join(lines)


def workbook_to_grids(workbook) -> dict[str, list[list[str]]]:
    out: dict[str, list[list[str]]] = {}
    for name in workbook.sheetnames:
        ws = workbook[name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        grid: list[list[str]] = []
        for r in range(1, max_row + 1):
            row_vals = []
            for c in range(1, max_col + 1):
                row_vals.append(_cell_str(ws.cell(row=r, column=c).value))
            grid.append(row_vals)
        out[name] = grid
    return out


def _col_letter_to_idx(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + ord(ch) - 64
    return n - 1


def _cell_ref_to_rc(ref: str) -> tuple[int, int]:
    m = re.match(r"([A-Z]+)(\d+)", ref)
    if not m:
        return 0, 0
    return int(m.group(2)) - 1, _col_letter_to_idx(m.group(1))


def workbook_to_grids_from_path(file_path: str) -> dict[str, list[list[str]]]:
    """优先 openpyxl；样式损坏的腾讯文档导出 xlsx 回退 zip+xml。"""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        try:
            return workbook_to_grids(wb)
        finally:
            wb.close()
    except Exception:
        return _workbook_to_grids_zip(file_path)


def _workbook_to_grids_zip(file_path: str) -> dict[str, list[list[str]]]:
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rel_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    with zipfile.ZipFile(file_path) as zf:
        strings: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            ss_root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in ss_root.findall(".//m:si", ns):
                texts = [t.text or "" for t in si.findall(".//m:t", ns)]
                strings.append("".join(texts))

        wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {r.attrib["Id"]: r.attrib["Target"] for r in rels_root}

        out: dict[str, list[list[str]]] = {}

        def read_sheet(target: str) -> list[list[str]]:
            path = "xl/" + target.lstrip("/")
            root = ET.fromstring(zf.read(path))
            rows: dict[int, dict[int, str]] = {}
            for cell in root.findall(".//m:sheetData/m:row/m:c", ns):
                ref = cell.attrib.get("r")
                if not ref:
                    continue
                r_idx, c_idx = _cell_ref_to_rc(ref)
                t = cell.attrib.get("t")
                v_el = cell.find("m:v", ns)
                if v_el is None or v_el.text is None:
                    val = ""
                elif t == "s":
                    val = strings[int(v_el.text)]
                else:
                    val = v_el.text
                rows.setdefault(r_idx, {})[c_idx] = _cell_str(val)
            if not rows:
                return []
            max_r = max(rows)
            grid: list[list[str]] = []
            for r in range(max_r + 1):
                row = rows.get(r, {})
                max_c = max(row) if row else -1
                grid.append([row.get(c, "") for c in range(max_c + 1)])
            return grid

        for sh in wb_root.findall(".//m:sheet", ns):
            name = sh.attrib.get("name") or "sheet"
            rid = sh.attrib.get(rel_ns)
            target = rid_to_target.get(rid or "")
            if target:
                out[name] = read_sheet(target)
        return out


def classify_sheet(grid: list[list[str]]) -> str:
    scan = grid[:HEADER_SCAN_ROWS]
    joined = "\n".join("\t".join(row) for row in scan)
    if "镜号" in joined and "画面" in joined:
        return "shot_table"
    total_chars = sum(len(_cell_str(c)) for row in grid for c in row)
    if total_chars > 200:
        return "outline"
    return "unknown"


def extract_outline_text(grid: list[list[str]]) -> str:
    parts: list[str] = []
    for row in grid:
        for cell in row:
            text = _cell_str(cell)
            if text:
                parts.append(text)
    return "\n\n".join(parts).strip()


def _header_keyword_hits(row: list[str]) -> int:
    joined = _norm_header("".join(row))
    hits = 0
    for kw in HEADER_KEYWORDS:
        if kw in joined:
            hits += 1
    return hits


def detect_header_row(grid: list[list[str]]) -> int | None:
    best_idx: int | None = None
    best_hits = 0
    for i, row in enumerate(grid[:HEADER_SCAN_ROWS]):
        hits = _header_keyword_hits(row)
        if hits >= HEADER_MIN_KEYWORD_HITS and hits >= best_hits:
            best_hits = hits
            best_idx = i
    return best_idx


def _match_field(header: str) -> str | None:
    norm = _norm_header(header)
    if not norm:
        return None
    for field, synonyms in FIELD_SYNONYMS.items():
        for syn in synonyms:
            syn_norm = _norm_header(syn)
            if norm == syn_norm or syn_norm in norm or norm in syn_norm:
                return field
    return None


def build_column_mapping(header_row: list[str]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    mapping: list[dict[str, Any]] = []
    unrecognized: list[dict[str, Any]] = []
    for col, header in enumerate(header_row):
        header_text = _cell_str(header)
        if not header_text:
            continue
        field = _match_field(header_text)
        if field:
            mapping.append(
                {
                    "col": col,
                    "header": header_text,
                    "field": field,
                    "confirmed": False,
                }
            )
        else:
            unrecognized.append(
                {
                    "col": col,
                    "header": header_text,
                    "sample": "",
                    "confirmed": False,
                }
            )
    return mapping, unrecognized


def _parse_shot_number(value: str) -> int | None:
    text = _cell_str(value)
    if not text:
        return None
    if re.fullmatch(r"\d+", text):
        return int(text)
    return None


def _parse_duration(value: str) -> float | None:
    text = _cell_str(value)
    if not text:
        return None
    try:
        num = float(text)
        if num > 0:
            return num
    except ValueError:
        pass
    return None


def _non_empty_cells(row: list[str]) -> list[tuple[int, str]]:
    return [(i, _cell_str(c)) for i, c in enumerate(row) if _cell_str(c)]


def _is_scene_marker_row(row: list[str], shot_col: int | None) -> tuple[bool, str]:
    cells = _non_empty_cells(row)
    if len(cells) == 0:
        return False, ""
    if len(cells) == 1:
        return True, cells[0][1]
    if shot_col is not None:
        shot_val = _cell_str(row[shot_col]) if shot_col < len(row) else ""
        if shot_val and _parse_shot_number(shot_val) is None:
            others = [c for i, c in cells if i != shot_col]
            if not others:
                return True, shot_val
    return False, ""


def _row_to_shot(
    row: list[str],
    column_mapping: list[dict[str, Any]],
    unrecognized_cols: list[dict[str, Any]],
) -> dict[str, Any]:
    shot: dict[str, Any] = {}
    extras: dict[str, str] = {}
    col_to_field = {m["col"]: m["field"] for m in column_mapping}
    for col, field in col_to_field.items():
        val = _cell_str(row[col]) if col < len(row) else ""
        if not val:
            continue
        if field == "shotNumber":
            num = _parse_shot_number(val)
            if num is not None:
                shot["shotNumber"] = num
        elif field == "duration":
            dur = _parse_duration(val)
            if dur is not None:
                shot["duration"] = dur
        elif field == "prompt":
            shot["prompt"] = val
            shot["description"] = val
        else:
            shot[field] = val
    for item in unrecognized_cols:
        col = item["col"]
        val = _cell_str(row[col]) if col < len(row) else ""
        if val:
            extras[item["header"]] = val
    if extras:
        shot["importExtras"] = extras
    return shot


def _make_segment_id() -> str:
    return f"seg-import-{uuid.uuid4().hex[:10]}"


def parse_shot_sheet(sheet_name: str, grid: list[list[str]]) -> dict[str, Any]:
    header_idx = detect_header_row(grid)
    if header_idx is None:
        return {
            "sheet_name": sheet_name,
            "kind": "shot_table",
            "error": "未找到表头行（需包含镜号、画面等关键词）",
        }

    header_row = grid[header_idx]
    column_mapping, unrecognized_columns = build_column_mapping(header_row)
    shot_col = next((m["col"] for m in column_mapping if m["field"] == "shotNumber"), None)

    scene_markers: list[dict[str, Any]] = []
    rows: list[dict[str, Any]] = []
    segments: list[dict[str, Any]] = []
    current_segment_id: str | None = None
    next_shot_after_marker: int | None = None

    for row_idx in range(header_idx + 1, len(grid)):
        row = grid[row_idx]
        if not any(_cell_str(c) for c in row):
            continue

        is_marker, marker_text = _is_scene_marker_row(row, shot_col)
        if is_marker:
            seg_id = _make_segment_id()
            segments.append({"id": seg_id, "title": marker_text, "description": "", "duration": 0})
            current_segment_id = seg_id
            if rows:
                next_shot_after_marker = rows[-1].get("shotNumber")
                if isinstance(next_shot_after_marker, int):
                    next_shot_after_marker += 1
            else:
                next_shot_after_marker = 1
            scene_markers.append(
                {
                    "row_index": row_idx,
                    "text": marker_text,
                    "applies_from_shot": next_shot_after_marker,
                    "segment_id": seg_id,
                    "confirmed": False,
                }
            )
            continue

        shot_num = None
        if shot_col is not None and shot_col < len(row):
            shot_num = _parse_shot_number(row[shot_col])
        if shot_num is None:
            continue

        parsed = _row_to_shot(row, column_mapping, unrecognized_columns)
        parsed["shotNumber"] = shot_num
        if current_segment_id:
            parsed["segmentId"] = current_segment_id
        if "duration" not in parsed:
            parsed["duration"] = 8
        rows.append(parsed)

    for item in unrecognized_columns:
        for row in rows:
            extras = row.get("importExtras") or {}
            if item["header"] in extras and not item.get("sample"):
                item["sample"] = extras[item["header"]][:120]

    return {
        "sheet_name": sheet_name,
        "kind": "shot_table",
        "header_row_index": header_idx,
        "column_mapping": column_mapping,
        "scene_markers": scene_markers,
        "segments": segments,
        "rows": rows,
        "rows_preview": rows[:8],
        "unrecognized_columns": unrecognized_columns,
        "stats": {
            "shot_count": len(rows),
            "marker_count": len(scene_markers),
        },
        "content_hash": None,
    }


def parse_outline_sheet(sheet_name: str, grid: list[list[str]]) -> dict[str, Any]:
    text = extract_outline_text(grid)
    return {
        "sheet_name": sheet_name,
        "kind": "outline",
        "text": text,
        "text_preview": text[:500] + ("…" if len(text) > 500 else ""),
        "stats": {"char_count": len(text)},
    }


PROMPT_EMPTY_THRESHOLD = 0.30


def self_check_shot_sheet(parsed: dict[str, Any]) -> dict[str, Any]:
    """解析后自检：镜号连续性、字段填充率、场次标记合理性。"""
    rows = parsed.get("rows") or []
    issues: list[dict[str, Any]] = []
    warnings: list[str] = []

    if not rows:
        return {
            "ok": False,
            "issues": [{"code": "no_rows", "message": "未解析到有效镜头行"}],
            "warnings": warnings,
            "stats": {"shot_count": 0, "prompt_empty_rate": 1.0},
        }

    numbers = sorted(int(r["shotNumber"]) for r in rows if r.get("shotNumber") is not None)
    duplicates = {n for n in numbers if numbers.count(n) > 1}
    if duplicates:
        issues.append(
            {
                "code": "duplicate_shot_numbers",
                "message": f"镜号重复: {sorted(duplicates)[:8]}",
            }
        )

    if numbers:
        expected = list(range(numbers[0], numbers[-1] + 1))
        missing = [n for n in expected if n not in numbers]
        if missing:
            issues.append(
                {
                    "code": "gap_shot_numbers",
                    "message": f"镜号不连续，缺失: {missing[:12]}{'…' if len(missing) > 12 else ''}",
                    "missing": missing[:50],
                }
            )

    empty_prompt = sum(
        1 for r in rows if not (r.get("prompt") or r.get("description") or "").strip()
    )
    empty_rate = empty_prompt / len(rows) if rows else 1.0
    if empty_rate > PROMPT_EMPTY_THRESHOLD:
        issues.append(
            {
                "code": "low_prompt_fill",
                "message": f"画面描述空行占比 {empty_rate:.0%}（阈值 {PROMPT_EMPTY_THRESHOLD:.0%}）",
            }
        )

    marker_count = parsed.get("stats", {}).get("marker_count", 0)
    if marker_count == 0 and len(rows) > 25:
        warnings.append("长分镜表未识别到场次标记，请确认 Excel 中场次行格式")

    unrec = parsed.get("unrecognized_columns") or []
    stats = {
        "shot_count": len(rows),
        "prompt_empty_rate": round(empty_rate, 3),
        "marker_count": marker_count,
        "unrecognized_column_count": len(unrec),
    }

    return {
        "ok": len(issues) == 0,
        "issues": issues,
        "warnings": warnings,
        "stats": stats,
    }


def attach_self_check(parsed: dict[str, Any]) -> dict[str, Any]:
    if parsed.get("kind") != "shot_table" or parsed.get("error"):
        return parsed
    parsed["self_check"] = self_check_shot_sheet(parsed)
    return parsed
